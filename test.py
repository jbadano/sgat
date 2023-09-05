import os
import sys
import logging
import json
import json_log_formatter
import csv
import requests
import oracledb
import openpyxl
import pyzipper
import zipfile
from urllib.parse import urlparse
from dotenv import load_dotenv
from google.cloud import pubsub_v1
from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
from openpyxl import Workbook


load_dotenv()
logger = logging.getLogger()
logHandler = logging.StreamHandler(sys.stdout)
logHandler.setFormatter(json_log_formatter.VerboseJSONFormatter())
logger.addHandler(logHandler)
logger.setLevel(logging.INFO)



def get_rm_source_file_fromsharepoint_tocsv(sharepoint_url, client_id, client_secret, remote_path, file_name):
    logger.info(msg=f'Inicio Extrae Archivo Sharepoint- {sharepoint_url} : {remote_path} : {file_name}')
    ctx = get_sharepoint_context_using_app(sharepoint_url, client_id, client_secret)
    relative_path = urlparse(sharepoint_url).path + "/" + remote_path + "/" + file_name

    with open(file_name, "wb") as local_file:
        file = ctx.web.get_file_by_server_relative_url(relative_path)
        file.download(local_file)
        ctx.execute_query()
    
    excel_file = file_name
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    csv_file = file_name.split(".")[0] + ".csv"
    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f, delimiter=';')
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
    return csv_file

def es_rut_valido(rutin, dvin):
    logger.info(msg=f'rutin: {rutin} dvin: {dvin}')
    if rutin == "":
        return False
    rutval = rutin.upper()
    dvval = dvin.upper()
    factors = [2, 3, 4, 5, 6, 7]
    partial_sum = 0
    factor_position = 0
    for digit in reversed(rutval):
        partial_sum += int(digit) * factors[factor_position]
        factor_position = (factor_position + 1) % 6
    verification_digit = (11 - partial_sum % 11) % 11
    if verification_digit < 10:
       verification_digit = str(verification_digit)
    else:
        verification_digit = 'K'
    logger.info(msg=f'verification_digit: {verification_digit}')
    return verification_digit == dvval
    
def obtener_pan_byrut(dbschema, rut, dv, panbin, panofuscut, cur):
    logger.info(msg=f'panbin: -{panbin}- panofuscut: -{panofuscut}-')    
    pantrx = "No Disponible"    
    panultimoout = "No Disponible"
    dbcentaltaout = "No Disponible"
    dbcuentaout = "No Disponible"
    dbrutout = "No Disponible"
    dbdvout = "No Disponible"
    querynobin = f'''SELECT SATPAN.TAR_CENTALTA, SATPAN.TAR_NUM_RUT, SATPAN.TAR_DIG_RUT, SATPAN.TAR_CUENTA_SAT, SATPAN.TAR_PAN, SATPAN.TAR_FECBLQ, SATPAN.TAR_TITULAR
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_NUM_RUT = '{rut}'
                AND TAR_DIG_RUT = '{dv}'
                AND TAR_TITULAR = 'T'
                AND (
                TAR_FECBLQ = '0001-01-01'
                OR 
                (TAR_FECBLQ = (
                SELECT MAX(TAR_FECBLQ)
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_NUM_RUT = '{rut}'
                AND TAR_DIG_RUT = '{dv}'
                AND TAR_TITULAR = 'T'
                )AND NOT EXISTS(
                SELECT 1
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_NUM_RUT = '{rut}'
                AND TAR_DIG_RUT = '{dv}'
                AND TAR_TITULAR = 'T'
                AND TAR_FECBLQ = '0001-01-01'
                )))
                ORDER BY SATPAN.TAR_FECBLQ ASC
            '''
    
    queryall = f'''SELECT SATPAN.TAR_CENTALTA, SATPAN.TAR_NUM_RUT, SATPAN.TAR_DIG_RUT, SATPAN.TAR_CUENTA_SAT, SATPAN.TAR_PAN, SATPAN.TAR_FECBLQ, SATPAN.TAR_TITULAR
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_NUM_RUT = '{rut}'
                AND TAR_DIG_RUT = '{dv}'
                AND TAR_TITULAR = 'T'
                AND (
                TAR_FECBLQ = '0001-01-01'
                OR 
                (TAR_FECBLQ = (
                SELECT MAX(TAR_FECBLQ)
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_NUM_RUT = '{rut}'
                AND TAR_DIG_RUT = '{dv}'
                AND TAR_TITULAR = 'T'
                )))
                ORDER BY SATPAN.TAR_FECBLQ DESC
            '''
    
    if panbin == "":
        logger.info(msg=f'querynobin: {querynobin}')     
        cur.execute(querynobin)
        rows = cur.fetchall()
        rowcount = len(rows)
        logger.info(msg=f'query cur.rowcount {rowcount}')    
        for TAR_CENTALTA, TAR_NUM_RUT, TAR_DIG_RUT, TAR_CUENTA_SAT, TAR_PAN, TAR_FECBLQ, TAR_TITULAR in rows:
            pantrx = TAR_PAN.strip()
            panultimoout = TAR_PAN.strip()
            dbcentaltaout = TAR_CENTALTA.strip()
            dbcuentaout = TAR_CUENTA_SAT.strip()
            dbrutout = TAR_NUM_RUT
            dbdvout = TAR_DIG_RUT.strip()
            break
    elif len(panofuscut) > 6: 
        logger.info(msg=f'queryall: {queryall}')     
        cur.execute(queryall)
        rows = cur.fetchall()
        rowcount = len(rows)
        logger.info(msg=f'query cur.rowcount {rowcount}')    
        for TAR_CENTALTA_ALL, TAR_NUM_RUT_ALL, TAR_DIG_RUT_ALL, TAR_CUENTA_SAT_ALL, TAR_PAN_ALL, TAR_FECBLQ_ALL, TAR_TITULAR_ALL in rows:
            pantrx_all = TAR_PAN_ALL.strip()
            panultimoout_all = TAR_PAN_ALL.strip()
            dbcentaltaout_all = TAR_CENTALTA_ALL.strip()
            dbcuentaout_all = TAR_CUENTA_SAT_ALL.strip()
            dbrutout_all = TAR_NUM_RUT_ALL
            dbdvout_all = TAR_DIG_RUT_ALL.strip()
            dbfecblq_all = TAR_FECBLQ_ALL.strip()
            pantrx = pantrx_all
            panultimoout = panultimoout_all
            dbcentaltaout = dbcentaltaout_all
            dbcuentaout = dbcuentaout_all
            dbrutout = dbrutout_all
            dbdvout = dbdvout_all
            if panofuscut == pantrx_all[0:6] + pantrx_all[-4:]:
                pantrx = pantrx_all
                panultimoout = panultimoout_all
                dbcentaltaout = dbcentaltaout_all
                dbcuentaout = dbcuentaout_all
                dbrutout = dbrutout_all
                dbdvout = dbdvout_all
                if dbfecblq_all != '0001-01-01':
                    for TAR_CENTALTA_ALL2, TAR_NUM_RUT_ALL2, TAR_DIG_RUT_ALL2, TAR_CUENTA_SAT_ALL2, TAR_PAN_ALL2, TAR_FECBLQ_ALL2, TAR_TITULAR_ALL2 in rows:                        
                        panultimoout_all2 = TAR_PAN_ALL2.strip()
                        dbcentaltaout_all2 = TAR_CENTALTA_ALL2.strip()
                        dbcuentaout_all2 = TAR_CUENTA_SAT_ALL2.strip()
                        dbrutout_all2 = TAR_NUM_RUT_ALL2
                        dbdvout_all2 = TAR_DIG_RUT_ALL2.strip()
                        dbfecblq_all2 = TAR_FECBLQ_ALL2.strip()
                        if dbcentaltaout_all == dbcentaltaout_all2 and dbcuentaout_all == dbcuentaout_all2 and dbfecblq_all2 == '0001-01-01':
                            panultimoout = panultimoout_all2
                            dbcentaltaout = dbcentaltaout_all2
                            dbcuentaout = dbcuentaout_all2
                            dbrutout = dbrutout_all2
                            dbdvout = dbdvout_all2
                            break
                break
            elif dbfecblq_all == '0001-01-01':
                pantrx = pantrx_all
                panultimoout = panultimoout_all
                dbcentaltaout = dbcentaltaout_all
                dbcuentaout = dbcuentaout_all
                dbrutout = dbrutout_all
                dbdvout = dbdvout_all
    else:
        logger.info(msg=f'queryall: {queryall}')     
        cur.execute(queryall)
        rows = cur.fetchall()
        rowcount = len(rows)
        logger.info(msg=f'query cur.rowcount {rowcount}')    
        for TAR_CENTALTA_ALL, TAR_NUM_RUT_ALL, TAR_DIG_RUT_ALL, TAR_CUENTA_SAT_ALL, TAR_PAN_ALL, TAR_FECBLQ_ALL, TAR_TITULAR_ALL in rows:
            pantrx_all = TAR_PAN_ALL.strip()
            panultimoout_all = TAR_PAN_ALL.strip()
            dbcentaltaout_all = TAR_CENTALTA_ALL.strip()
            dbcuentaout_all = TAR_CUENTA_SAT_ALL.strip()
            dbrutout_all = TAR_NUM_RUT_ALL
            dbdvout_all = TAR_DIG_RUT_ALL.strip()
            dbfecblq_all = TAR_FECBLQ_ALL.strip()
            pantrx = pantrx_all
            panultimoout = panultimoout_all
            dbcentaltaout = dbcentaltaout_all
            dbcuentaout = dbcuentaout_all
            dbrutout = dbrutout_all
            dbdvout = dbdvout_all
            if panbin == pantrx_all[0:6]:
                pantrx = pantrx_all
                panultimoout = panultimoout_all
                dbcentaltaout = dbcentaltaout_all
                dbcuentaout = dbcuentaout_all
                dbrutout = dbrutout_all
                dbdvout = dbdvout_all
                if dbfecblq_all != '0001-01-01':
                    for TAR_CENTALTA_ALL2, TAR_NUM_RUT_ALL2, TAR_DIG_RUT_ALL2, TAR_CUENTA_SAT_ALL2, TAR_PAN_ALL2, TAR_FECBLQ_ALL2, TAR_TITULAR_ALL2 in rows:                        
                        panultimoout_all2 = TAR_PAN_ALL2.strip()
                        dbcentaltaout_all2 = TAR_CENTALTA_ALL2.strip()
                        dbcuentaout_all2 = TAR_CUENTA_SAT_ALL2.strip()
                        dbrutout_all2 = TAR_NUM_RUT_ALL2
                        dbdvout_all2 = TAR_DIG_RUT_ALL2.strip()
                        dbfecblq_all2 = TAR_FECBLQ_ALL2.strip()
                        if dbcentaltaout_all == dbcentaltaout_all2 and dbcuentaout_all == dbcuentaout_all2 and dbfecblq_all2 == '0001-01-01':
                            panultimoout = panultimoout_all2
                            dbcentaltaout = dbcentaltaout_all2
                            dbcuentaout = dbcuentaout_all2
                            dbrutout = dbrutout_all2
                            dbdvout = dbdvout_all2
                            break
                break
            elif dbfecblq_all == '0001-01-01':
                pantrx = pantrx_all
                panultimoout = panultimoout_all
                dbcentaltaout = dbcentaltaout_all
                dbcuentaout = dbcuentaout_all
                dbrutout = dbrutout_all
                dbdvout = dbdvout_all
    return (pantrx,panultimoout,dbrutout,dbdvout,dbcentaltaout,dbcuentaout)

def obtener_pan_by12(dbschema, numaut, codcom, imptrn, fectrn, rut, dv, panbin, panofuscut, cur):
    pantrx = "No Disponible"
    dbcentalta = "No Disponible"
    dbcuenta = "No Disponible"
    panultimoout = "No Disponible"
    dbcentaltaout = "No Disponible"
    dbcuentaout = "No Disponible"
    dbrutout = "No Disponible"
    dbdvout = "No Disponible"
    query = f'''SELECT SATPAN.TAR_CENTALTA, SATPAN.TAR_NUM_RUT, SATPAN.TAR_DIG_RUT, SATPAN.TAR_CUENTA_SAT, SATPAN.TAR_PAN, SATPAN.TAR_FECBLQ, SATPAN.TAR_TITULAR
                        FROM {dbschema}.MPDT012 M012, {dbschema}.PRECON_SATIF_PAN SATPAN
                        WHERE M012.numaut = '{numaut}'
                        and M012.codcom = '{codcom}'
                        and M012.impfac = {imptrn}
                        and M012.fecfac = '{fectrn}'
                        and SATPAN.TAR_PAN = M012.PAN'''
    logger.info(msg=f'query: {query}')      
    cur.execute(query)
    rows = cur.fetchall()
    rowcount = len(rows)
    logger.info(msg=f'query cur.rowcount {rowcount}')    
    findult = False
    for TAR_CENTALTA, TAR_NUM_RUT, TAR_DIG_RUT, TAR_CUENTA_SAT, TAR_PAN, TAR_FECBLQ, TAR_TITULAR in rows:                    
        pantrx = TAR_PAN.strip()        
        dbcentalta = TAR_CENTALTA.strip()
        dbcuenta = TAR_CUENTA_SAT.strip()
        dbrut = TAR_NUM_RUT
        dbdv = TAR_DIG_RUT.strip()
        dbfecblq = TAR_FECBLQ.strip()
        if dbfecblq == '0001-01-01':
            panultimoout = TAR_PAN.strip()
            dbrutout = dbrut
            dbdvout = dbdv
            dbcentaltaout = dbcentalta
            dbcuentaout = dbcuenta
        else:
            findult = True
        
    if findult:    
        responseCuenta = obtener_ult_pan_cuenta(dbschema,dbcentalta, dbcuenta, cur)
        panultimoout = responseCuenta[0]
        dbrutout = responseCuenta[1]
        dbdvout = responseCuenta[2]
        dbcentaltaout = responseCuenta[3]
        dbcuentaout = responseCuenta[4]

    return (pantrx,panultimoout,dbrutout,dbdvout,dbcentaltaout,dbcuentaout)
        

    
def obtener_pan_by04(dbschema, numaut, codcom, imptrn, fectrn, rut, dv, panbin, panofuscut, cur):
    pantrx = "No Disponible"
    dbcentalta = "No Disponible"
    dbcuenta = "No Disponible"
    panultimoout = "No Disponible"
    dbcentaltaout = "No Disponible"
    dbcuentaout = "No Disponible"
    dbrutout = "No Disponible"
    dbdvout = "No Disponible"

    query = f'''SELECT SATPAN.TAR_CENTALTA, SATPAN.TAR_NUM_RUT, SATPAN.TAR_DIG_RUT, SATPAN.TAR_CUENTA_SAT, SATPAN.TAR_PAN, SATPAN.TAR_FECBLQ, SATPAN.TAR_TITULAR
                        FROM {dbschema}.MPDT004 M004, {dbschema}.PRECON_SATIF_PAN SATPAN
                        WHERE M004.numaut = '{numaut}'
                        and M004.codcom = '{codcom}'
                        and M004.imptrn = {imptrn}
                        and M004.fectrn = '{fectrn}'
                        and SATPAN.TAR_PAN = M004.PAN'''
    logger.info(msg=f'query: {query}')      
    cur.execute(query)
    rows = cur.fetchall()
    rowcount = len(rows)
    logger.info(msg=f'query cur.rowcount {rowcount}')    
    findult = False
    for TAR_CENTALTA, TAR_NUM_RUT, TAR_DIG_RUT, TAR_CUENTA_SAT, TAR_PAN, TAR_FECBLQ, TAR_TITULAR in rows:                    
        pantrx = TAR_PAN.strip()        
        dbcentalta = TAR_CENTALTA.strip()
        dbcuenta = TAR_CUENTA_SAT.strip()
        dbrut = TAR_NUM_RUT
        dbdv = TAR_DIG_RUT.strip()
        dbfecblq = TAR_FECBLQ.strip()
        if dbfecblq == '0001-01-01':
            panultimoout = TAR_PAN.strip()
            dbrutout = dbrut
            dbdvout = dbdv
            dbcentaltaout = dbcentalta
            dbcuentaout = dbcuenta
        else:
            findult = True
        
    if findult:    
        responseCuenta = obtener_ult_pan_cuenta(dbschema,dbcentalta, dbcuenta, cur)
        panultimoout = responseCuenta[0]
        dbrutout = responseCuenta[1]
        dbdvout = responseCuenta[2]
        dbcentaltaout = responseCuenta[3]
        dbcuentaout = responseCuenta[4]

    return (pantrx,panultimoout,dbrutout,dbdvout,dbcentaltaout,dbcuentaout)


def ret_query_aplicarmsinnumaut(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):
    if codcom.lstrip('0') != '10001827':
        query_aplicarmsinnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom}'
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        query_aplicarmsinnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'                           
            AND T855.IMPOPER = {imptrn}
        '''
    return query_aplicarmsinnumaut

def ret_query_aplicarmconnumaut(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):
    if codcom.lstrip('0') != '10001827':
        query_aplicarmconnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom}'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        query_aplicarmconnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    return query_aplicarmconnumaut

def ret_querysatclpr_aplicarmsinnumaut(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):
    if codcom.lstrip('0') != '10001827':
        querysatclpr_aplicarmsinnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom.lstrip('0')}'                            
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        querysatclpr_aplicarmsinnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'                         
            AND T855.IMPOPER = {imptrn}
        '''
    return querysatclpr_aplicarmsinnumaut

def ret_querysatclpr_aplicarmconnumaut(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):   
    if codcom.lstrip('0') != '10001827':
        querysatclpr_aplicarmconnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom.lstrip('0')}'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        querysatclpr_aplicarmconnumaut = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'
            AND (T855.PAN = '{panquery}   ' OR T855.PAN = '{panquery[0:6]}000000{panquery[-4:]}   ' OR T855.PAN = '                   ' OR T855.PAN = '{panofus}   ')
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    return querysatclpr_aplicarmconnumaut

def ret_querysatclpr_aplicarmsinpan(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):   
    if codcom.lstrip('0') != '10001827':
        querysatclpr_aplicarmsinpan = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'            
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom.lstrip('0')}'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        querysatclpr_aplicarmsinpan = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM SATCLPR.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'            
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    return querysatclpr_aplicarmsinpan
        
def ret_query_aplicarmsinpan(dbschema,imptrn,numaut,fectrn,panquery,panofus,codcom):   
    if codcom.lstrip('0') != '10001827':
        query_aplicarmsinpan = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'            
            AND T855.CODENT = '0015'
            AND T855.CODCOM = '{codcom}'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    else:
        query_aplicarmsinpan = f'''
            SELECT LPAD(T855.INDERROR,3,0) as INDERROR, TO_CHAR(T221.CATDATOS) as CATDATOS, SATIF_PAN.TAR_CUENTA_SAT as TARCUENTA, T012.numaut as NUMAUT12
            FROM {dbschema}.MPDT855 T855
            LEFT JOIN {dbschema}.MPDT221_SAT T221
            ON T221.CODCATALO = 'MERCHANTRECHAZ'
            AND TRIM(T221.CODELEMEN) = T855.CODRECHAZO
            LEFT JOIN {dbschema}.PRECON_SATIF_PAN SATIF_PAN
            ON SATIF_PAN.TAR_PAN = '{panquery}'
            LEFT JOIN {dbschema}.MPDT012  T012
            ON T012.CUENTA = SATIF_PAN.TAR_CUENTA_SAT
            AND T012.CENTALTA = SATIF_PAN.TAR_CENTALTA
            AND T012.CODENT = SATIF_PAN.TAR_CODENT
            AND T012.IMPFAC = {imptrn}
            AND T012.NUMAUT = '{numaut}'
            WHERE T855.FECOPER = '{fectrn}'            
            AND T855.CODENT = '0015'
            AND T855.TIPFRAN = '1399'
            AND T855.NUMAUT = '{numaut}'
            AND T855.IMPOPER = {imptrn}
        '''
    return query_aplicarmsinpan

def obtener_ult_pan_cuenta(dbschema, dbcentalta, dbcuenta, cur):
    query_ult_pan = f'''SELECT SATPAN.TAR_CENTALTA, SATPAN.TAR_NUM_RUT, SATPAN.TAR_DIG_RUT, SATPAN.TAR_CUENTA_SAT, SATPAN.TAR_PAN
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_CODENT = '0015'
                AND TAR_CENTALTA = '{dbcentalta}'
                AND TAR_CUENTA_SAT = '{dbcuenta}'
                AND TAR_TITULAR = 'T'
                AND (
                TAR_FECBLQ = '0001-01-01'
                OR 
                (TAR_FECBLQ = (
                SELECT MAX(TAR_FECBLQ)
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_CODENT = '0015'
                AND TAR_CENTALTA = '{dbcentalta}'
                AND TAR_CUENTA_SAT = '{dbcuenta}'
                AND TAR_TITULAR = 'T'
                )
                AND NOT EXISTS(
                SELECT 1
                from {dbschema}.PRECON_SATIF_PAN SATPAN
                WHERE TAR_CODENT = '0015'
                AND TAR_CENTALTA = '{dbcentalta}'
                AND TAR_CUENTA_SAT = '{dbcuenta}'
                AND TAR_TITULAR = 'T'
                AND TAR_FECBLQ = '0001-01-01'
                )))'''    
    logger.info(msg=f'query: {query_ult_pan}')      
    cur.execute(query_ult_pan)
    rows = cur.fetchall()
    rowcount = len(rows)
    logger.info(msg=f'query cur.rowcount {rowcount}')
    if rowcount == 1:
        for TAR_CENTALTA, TAR_NUM_RUT, TAR_DIG_RUT, TAR_CUENTA_SAT, TAR_PAN in rows:                    
            pandes = TAR_PAN
            rutout = TAR_NUM_RUT
            dvout = TAR_DIG_RUT
            centaltaout = TAR_CENTALTA
            cuentaout = TAR_CUENTA_SAT
    else:
        pandes = "No Disponible"
        rutout = "No Disponible"
        dvout = "No Disponible"
        centaltaout = "No Disponible"
        cuentaout = "No Disponible"
    return (pandes, rutout, dvout, centaltaout, cuentaout)

def valida_aplicarm(dbschema,codcom,imptrn,fectrn,numaut,numautprevio,panofuscut,pantrx,panofus,validbin,satfcc,trx,cur):
    aplicarm = "NO APLICA RM"
    if codcom == '000000010000202':
        aplicarm = "APLICA RM"
    elif satfcc == "SATIF":
        aplicarm = "NO APLICA RM SATIF"
    elif satfcc == "FCC":
        aplicarm = "NO APLICA RM FCC"
    else:        
        
        encontrado855 = False
        
        query_aplicarmconnumaut = ret_query_aplicarmconnumaut(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
        logger.info(msg=f'query_aplicarm: {query_aplicarmconnumaut}')      
        cur.execute(query_aplicarmconnumaut)
        rows = cur.fetchall()
        rowcount = len(rows)
        if rowcount < 1:
            query_aplicarmsinnumaut = ret_query_aplicarmsinnumaut(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
            logger.info(msg=f'query_aplicarm: {query_aplicarmsinnumaut}')      
            cur.execute(query_aplicarmsinnumaut)
            rows = cur.fetchall()
            rowcount = len(rows)
        if rowcount < 1:
            querysatclpr_aplicarmconnumaut = ret_querysatclpr_aplicarmconnumaut(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
            logger.info(msg=f'query_aplicarm: {querysatclpr_aplicarmconnumaut}')      
            cur.execute(querysatclpr_aplicarmconnumaut)
            rows = cur.fetchall()
            rowcount = len(rows)
        if rowcount < 1:
            querysatclpr_aplicarmsinnumaut = ret_querysatclpr_aplicarmsinnumaut(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
            logger.info(msg=f'query_aplicarm: {querysatclpr_aplicarmsinnumaut}')      
            cur.execute(querysatclpr_aplicarmsinnumaut)
            rows = cur.fetchall()
            rowcount = len(rows)
        
        if rowcount < 1:
            querysatclpr_aplicarmsinpan = ret_querysatclpr_aplicarmsinpan(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
            logger.info(msg=f'query_aplicarm: {querysatclpr_aplicarmsinpan}')      
            cur.execute(querysatclpr_aplicarmsinpan)
            rows = cur.fetchall()
            rowcount = len(rows)

        if rowcount < 1:
            query_aplicarmsinpan = ret_query_aplicarmsinpan(dbschema,imptrn,numaut,fectrn,pantrx,panofus,codcom)
            logger.info(msg=f'query_aplicarm: {query_aplicarmsinpan}')      
            cur.execute(query_aplicarmsinpan)
            rows = cur.fetchall()
            rowcount = len(rows)

        logger.info(msg=f'query cur.rowcount {rowcount}')

        if rowcount > 0:
            encontrado855 = True

        if encontrado855:

            for INDERROR, CATDATOS, TARCUENTA, NUMAUT12 in rows:
                if CATDATOS is None:
                    aplicarm = "APLICA RM"
                elif CATDATOS.strip() != '333':
                    aplicarm = "APLICA RM"

                if NUMAUT12 is None:
                    aplicarm = "APLICA RM"
                else:
                    aplicarm = "NO APLICA RM 12"
                    
            if trx == '217' or trx == '238' or trx == '0217' or trx == '0238':
                query12pag = f'''
                select T012.NUMAUT as NUMAUT, T012.IMPFAC as IMPFAC, T012.TIPOFAC as TIPOFAC
                FROM {dbschema}.PRECON_SATIF_PAN SATIF_PAN
                LEFT JOIN {dbschema}.MPDT012  T012
                ON SATIF_PAN.TAR_CUENTA_SAT = T012.CUENTA
                AND SATIF_PAN.TAR_CENTALTA = T012.CENTALTA
                AND SATIF_PAN.TAR_CODENT = T012.CODENT
                AND T012.NUMAUT = '{numaut}'
                AND T012.IMPFAC = {imptrn}
                AND TIPOFAC in ('6004','3005')
                WHERE TAR_PAN = '{pantrx}'            
                '''
                logger.info(msg=f'query_aplicarm: {query12pag}')      
                cur.execute(query12pag)
                rows = cur.fetchall()
                rowcount = len(rows)
                logger.info(msg=f'query cur.rowcount {rowcount}')
                for NUMAUT, IMPFAC, TIPOFAC in rows:
                    if IMPFAC is not None:
                        aplicarm = "APLICA RM3005"
            
            if trx == '400' or trx == '0400':

                if numautprevio == '000000':
                    numautprevio = numaut

                aplicarm = "NO APLICA RM 004"
                if codcom.lstrip('0') != '10001827':
                    query400 = f'''SELECT M004.NUMAUT as NUMAUT
                            FROM {dbschema}.MPDT004 M004
                            WHERE M004.numaut = '{numautprevio}'
                            and M004.codcom = '{codcom}'
                            and M004.imptrn = {imptrn}
                            and M004.fectrn = '{fectrn}'   
                            and M004.PAN = '{pantrx}'            
                    '''
                else:
                    query400 = f'''SELECT M004.NUMAUT as NUMAUT
                            FROM {dbschema}.MPDT004 M004
                            WHERE M004.numaut = '{numautprevio}'
                            and M004.tipfran = '1399'
                            and M004.imptrn = {imptrn}
                            and M004.fectrn = '{fectrn}'   
                            and M004.PAN = '{pantrx}'            
                    '''

                logger.info(msg=f'query_aplicarm: {query400}')      
                cur.execute(query400)
                rows = cur.fetchall()
                rowcount = len(rows)
                logger.info(msg=f'query cur.rowcount {rowcount}')
                for NUMAUT in rows:
                    if NUMAUT is not None:
                        aplicarm = "APLICA RM"
                    else:
                        aplicarm = "NO APLICA RM 004"
                
            if trx == '401' or trx == '0401' or trx == '402' or trx == '0402':
                aplicarm = "NO APLICA RM 012"
                query12 = f'''SELECT M012.NUMAUT as NUMAUT
                        FROM {dbschema}.MPDT012 M012
                        WHERE M012.numaut = '{numaut}'
                        and M012.codcom = '{codcom}'
                        and M012.impfac = {imptrn}
                        and M012.PAN = '{pantrx}'
                        '''
                logger.info(msg=f'query_aplicarm: {query12}')      
                cur.execute(query12)
                rows = cur.fetchall()
                rowcount = len(rows)
                logger.info(msg=f'query cur.rowcount {rowcount}')
                for NUMAUT in rows:
                    if NUMAUT is not None:
                        aplicarm = "APLICA RM"
                    else:
                        aplicarm = "NO APLICA RM 012"
        else:
            aplicarm = "NO APLICA RM 855"
        
        if aplicarm == "NO APLICA RM 855" and pantrx[0:6] != panofus[0:6] and pantrx[-4:]!=panofus[-4:]:
            aplicarm = "NO APLICA RM INCONSISTENTE"

    return aplicarm

def valida_satfcc(rutin, dvin, centaltain, cuentain, dbschema, cur):

    satfcc = "SAT"

    if cuentain.startswith('8'):
        satfcc = "SATIF"

    query_satfcc = f"""SELECT CUENTA_FCC FROM {dbschema}.CARTERA_CGDA_FCC
                    WHERE RUT_CLI = {rutin}
                    AND UPPER(DIG_VER) = UPPER('{dvin}') 
                    AND CONTRATO = '0015{centaltain}{cuentain}'
                    """
    
    logger.info(msg=f'query_satfcc: {query_satfcc}')     
    cur.execute(query_satfcc)
    rows = cur.fetchall()
    rowcount = len(rows)
    logger.info(msg=f'query cur.rowcount {rowcount}')
    for CUENTA_FCC in rows:
        if CUENTA_FCC is not None:
            satfcc = "FCC"
    
    return satfcc



def rm_complete_data(dbhost, dbuser, dbpass, dbport, dbservice, dbschema, local_file_path, pass_file):
    logger.info(msg='Inicio rm_complete_data')
    oracledb.init_oracle_client()
    params = oracledb.ConnectParams(host=dbhost, port=dbport, service_name=dbservice)
    con = oracledb.connect(user=dbuser, password=dbpass, params=params)
    cur = con.cursor()

    queryselectbin = f"""
                    SELECT CODELEMEN FROM {dbschema}.MPDT221_SAT 
                    WHERE CODCATALO IN ('BINESMASTER','BINESVISA')
                    """
    
    logger.info(msg=f'querynobin: {queryselectbin}')     
    cur.execute(queryselectbin)
    rows = cur.fetchall()
    rowcount = len(rows)
    logger.info(msg=f'query cur.rowcount {rowcount}')    
    bines = ""
    for CODELEMEN in rows:
        bines = bines + CODELEMEN[0].strip() + ","
        
    

    output_file = local_file_path.split(".")[0]+"OUT.csv"
    with open(local_file_path, 'r') as archivo_csv, open(output_file, 'w', newline='') as archivo_salida:
        lector_csv = csv.DictReader(archivo_csv, delimiter=";")
        columnas_originales = lector_csv.fieldnames        
        columnas_nuevas = columnas_originales + ["PANTRX"] + ["PANULT"] + ["RUTOUT"] + ["DVOUT"] + ["CENTALTAOUT"] + ["CUENTAOUT"] + ["SATFCC"] + ["APLICARM"]
        escritor_csv = csv.DictWriter(archivo_salida, fieldnames=columnas_nuevas)
        escritor_csv.writeheader()
        for fila in lector_csv:
            try:
                pantrx = "No Disponible"
                panult = "No Disponible"
                rutout = "No Disponible"
                dvout = "No Disponible"
                centaltaout = "No Disponible"
                cuentaout = "No Disponible"
                satfcc = "No Disponible"
                rut = fila["NumeroDocumentoIdentificacion"]
                dv = fila["Digito"]
                dv = dv.strip()
                rut = rut.strip()
                rut = rut.lstrip('0')
                panofus = fila["PAN"]
                panofus = panofus.strip()
                panbin = panofuscut = panofus[0:6]
                panofuscut = panofus[0:6] + panofus[-4:]

                codcom = fila["CodigoComercio"].strip()
                codcom = codcom.zfill(15)

                imptrn = fila["MontoTX"].strip()
                
                numaut = fila["CodigoAutorizacion"].strip()
                numaut = numaut[-6:]

                fectrn = fila["Fecha_C"].strip()
                fectrn = '-'.join([fectrn[:4], fectrn[4:6], fectrn[6:]])

                trx = fila["trx"].strip()    

                numautprevio = fila["CodigoAutorizacionPrevio"].strip()
                numautprevio = numautprevio[-6:]            
                
                validbin =False
                if panbin in bines:
                    validbin = True


                if trx == '217' or trx == '238' or trx == '0217' or trx == '0238':
                    response_find = obtener_pan_by12(dbschema, numaut, codcom, imptrn, fectrn, rut, dv, panbin, panofuscut, cur)
                    if response_find[0] == "No Disponible" and es_rut_valido(rut, dv):
                        response_find = obtener_pan_byrut(dbschema, rut, dv, panbin, panofuscut, cur)
                else:
                    response_find = obtener_pan_by04(dbschema, numaut, codcom, imptrn, fectrn, rut, dv, panbin, panofuscut, cur)
                    if response_find[0] == "No Disponible" and es_rut_valido(rut, dv):
                        response_find = obtener_pan_byrut(dbschema, rut, dv, panbin, panofuscut, cur)

                pantrx = response_find[0]
                panult = response_find[1]
                rutout = response_find[2]
                dvout = response_find[3]
                centaltaout = response_find[4]
                cuentaout = response_find[5]

                if cuentaout != "No Disponible":
                    satfcc = valida_satfcc(rutout, dvout ,centaltaout, cuentaout, dbschema, cur)
                
                aplicarm = "No Disponible"
                if pantrx != "No Disponible":
                    aplicarm = valida_aplicarm(dbschema,codcom,imptrn,fectrn,numaut,numautprevio,panofuscut, pantrx, panofus,validbin,satfcc, trx,cur)

                filaout = dict(fila)
                filaout["PANTRX"] = str(pantrx).strip()
                filaout["PANULT"] = str(panult).strip()
                filaout["RUTOUT"] = str(rutout).strip()
                filaout["DVOUT"] = str(dvout).strip()
                filaout["CENTALTAOUT"] = str(centaltaout).strip()
                filaout["CUENTAOUT"] = str(cuentaout).strip()
                filaout["SATFCC"] = str(satfcc).strip()
                filaout["APLICARM"] = aplicarm
                escritor_csv.writerow(filaout)
            except Exception as e:
                logger.error(msg='Error en rm_complete_data',exc_info=True)
                filaout = dict(fila)
                filaout["PANTRX"] = "ERROR"
                filaout["PANULT"] = "ERROR"
                filaout["RUTOUT"] = "ERROR"
                filaout["DVOUT"] = "ERROR"
                filaout["CENTALTAOUT"] = "ERROR"
                filaout["CUENTAOUT"] = "ERROR"
                filaout["SATFCC"] = "ERROR"
                filaout["APLICARM"] = "ERROR"
                escritor_csv.writerow(filaout)                
                continue
    cur.close()
    con.close()   

    archivo_origen_csv = output_file
    output_file_xlsx = local_file_path.split(".")[0]+"OUT.xlsx"
    with open(archivo_origen_csv, 'r') as csv_file_fin:
        csv_reader_fin = csv.reader(csv_file_fin)
        datos_csv_fin = list(csv_reader_fin)

    libro_trabajo = Workbook()
    hoja = libro_trabajo.active
    for fila_fin in datos_csv_fin:
        hoja.append(fila_fin)
    libro_trabajo.save(output_file_xlsx)

    archivo_origen = output_file_xlsx
    archivo_destino = output_file_xlsx.replace(".xlsx",".zip")
    contraseña = pass_file
    contraseña_bytes = contraseña.encode('utf-8') 
    with pyzipper.AESZipFile(archivo_destino, 'w', compression=zipfile.ZIP_LZMA) as zf:
        zf.setpassword(contraseña_bytes)
        zf.setencryption(pyzipper.WZ_AES, nbits=128)
        zf.write(archivo_origen)
    return archivo_destino
cod_trx_switcher = {
    "200":"1001",
    "217":"1068",
    "400":"1068",
    "220":"1047",
    "236":"1005",
    "238":"1068",
    "402":"6018",
    "274":"2003",
    "401":"6018",
    "403":"3005",
    "473":"3005"
}

def out_code_trx(cod_trx):    
    return cod_trx_switcher.get(cod_trx,"ERROR")

def rm_generate_rm_file(localfile_path, pass_file):
    logger.info(msg='Inicio rm_generate_rm_file')
        
    in_file = localfile_path.split(".")[0]+"OUT.csv"
    out_file = localfile_path.split(".")[0]+"OUT-RM.csv"
    with open(in_file, 'r') as archivo_csv, open(out_file, 'w', newline='') as archivo_salida:
        lector_csv = csv.DictReader(archivo_csv, delimiter=",")
        columnas_originales = lector_csv.fieldnames        
        columnas_nuevas = ["COD_COM"] + ["COD_TRX"] + ["FEC_TRX"] + ["HOR_TRX"] + ["LOCAL_COM"] + ["TERMINAL_COM"] + ["RUT"] + ["DV_RUT"] + ["TARJETA"] + ["CUOTAS"] + ["MONTO"] + ["TIPO_VENTA"] + ["ABONO"] + ["TIPO_DOCTO"] + ["NUM_DOCTO"] + ["COD_AUTORIZA"] + ["COMPRA_CUOTAS"] + ["PLAZO_PAGO_DIF"] + ["COD_AUTORIZA_PREV"] + ["MODALIDAD"] + ["FEC_ARCH_NEGATIVO"] + ["NUMEXTCTA"] + ["NUMMOVEXT"]
        escritor_csv = csv.DictWriter(archivo_salida, fieldnames=columnas_nuevas)
        escritor_csv.writeheader()
        for fila in lector_csv:
            try:               
                aplicarm = fila["APLICARM"]          
                if aplicarm == "APLICA RM" or aplicarm == "APLICA RM3005":
                    filaout = dict()
                    COD_COM = fila["CodigoComercio"].strip()
                    COD_COM = COD_COM.zfill(15)
                    filaout["COD_COM"] = COD_COM
                    COD_TRX = fila["trx"].strip().lstrip('0')
                    COD_TRX = out_code_trx(COD_TRX)
                    filaout["COD_TRX"] = COD_TRX
                    FEC_TRX = fila["Fecha_C"].strip()                    
                    filaout["FEC_TRX"] = FEC_TRX
                    HOR_TRX = fila["Hora"].strip()
                    filaout["HOR_TRX"] = HOR_TRX
                    LOCAL_COM = fila["Local"].strip()
                    LOCAL_COM = LOCAL_COM[-4:].zfill(4)
                    filaout["LOCAL_COM"] = LOCAL_COM
                    TERMINAL_COM = fila["Terminal"].strip()
                    TERMINAL_COM = TERMINAL_COM[-4:].zfill(4)                    
                    filaout["TERMINAL_COM"] = TERMINAL_COM
                    RUT = fila["RUTOUT"].strip()
                    filaout["RUT"] = RUT
                    DV_RUT = fila["DVOUT"].strip()
                    filaout["DV_RUT"] = DV_RUT
                    TARJETA = fila["PANULT"].strip()
                    filaout["TARJETA"] = TARJETA
                    CUOTAS = fila["Cuotas"].strip().lstrip('0')
                    CUOTAS = CUOTAS.zfill(2) 
                    filaout["CUOTAS"] = CUOTAS
                    MONTO = fila["MontoTX"].strip()
                    MONTO = MONTO.zfill(9)
                    filaout["MONTO"] = MONTO
                    TIPO_VENTA = fila["MododePago"].strip()
                    filaout["TIPO_VENTA"] = TIPO_VENTA
                    ABONO = fila["OtroMonto"].strip()
                    ABONO = ABONO.zfill(8)
                    filaout["ABONO"] = ABONO
                    TIPO_DOCTO = fila["TipoDocumentoVenta"].strip()
                    filaout["TIPO_DOCTO"] = TIPO_DOCTO
                    NUM_DOCTO = fila["N°DocumentoVenta"].strip()
                    NUM_DOCTO = NUM_DOCTO[-8:].zfill(8)
                    filaout["NUM_DOCTO"] = NUM_DOCTO
                    COD_AUTORIZA = fila["CodigoAutorizacion"].strip()
                    COD_AUTORIZA = COD_AUTORIZA.zfill(12)
                    filaout["COD_AUTORIZA"] = COD_AUTORIZA
                    COMPRA_CUOTAS = fila["Indicadorcompraencuotas"].strip()
                    filaout["COMPRA_CUOTAS"] = COMPRA_CUOTAS
                    PLAZO_PAGO_DIF = fila["PlazodePagodiferido"].strip()
                    PLAZO_PAGO_DIF = PLAZO_PAGO_DIF[-2:].zfill(2)
                    filaout["PLAZO_PAGO_DIF"] = PLAZO_PAGO_DIF
                    COD_AUTORIZA_PREV = fila["CodigoAutorizacionPrevio"].strip()
                    COD_AUTORIZA_PREV = COD_AUTORIZA_PREV[-12:].zfill(12)
                    filaout["COD_AUTORIZA_PREV"] = COD_AUTORIZA_PREV
                    MODALIDAD = fila["Modalidad"].strip()
                    filaout["MODALIDAD"] = MODALIDAD
                    FEC_ARCH_NEGATIVO = fila["FechaNegativo"].strip()
                    FEC_ARCH_NEGATIVO = FEC_ARCH_NEGATIVO.zfill(8)
                    filaout["FEC_ARCH_NEGATIVO"] = FEC_ARCH_NEGATIVO
                    
                    NUMEXTCTA = ""
                    filaout["NUMEXTCTA"] = NUMEXTCTA
                    NUMMOVEXT = ""
                    filaout["NUMMOVEXT"] = NUMMOVEXT
                    escritor_csv.writerow(filaout)
            except Exception as e:
                logger.error(msg='Error en rm_complete_data',exc_info=True)                
                filaout = dict()
                filaout["COD_COM"] = "ERROR"
                filaout["COD_TRX"] = "ERROR"
                filaout["FEC_TRX"] = "ERROR"
                filaout["HOR_TRX"] = "ERROR"
                filaout["LOCAL_COM"] = "ERROR"
                filaout["TERMINAL_COM"] = "ERROR"
                filaout["RUT"] = "ERROR"
                filaout["DV_RUT"] = "ERROR"
                filaout["TARJETA"] = "ERROR"
                filaout["CUOTAS"] = "ERROR"
                filaout["MONTO"] = "ERROR"
                filaout["TIPO_VENTA"] = "ERROR"
                filaout["ABONO"] = "ERROR"
                filaout["TIPO_DOCTO"] = "ERROR"
                filaout["NUM_DOCTO"] = "ERROR"
                filaout["COD_AUTORIZA"] = "ERROR"
                filaout["COMPRA_CUOTAS"] = "ERROR"
                filaout["PLAZO_PAGO_DIF"] = "ERROR"
                filaout["COD_AUTORIZA_PREV"] = "ERROR"
                filaout["MODALIDAD"] = "ERROR"
                filaout["FEC_ARCH_NEGATIVO"] = "ERROR"
                filaout["NUMEXTCTA"] = "ERROR"
                filaout["NUMMOVEXT"] = "ERROR"
                escritor_csv.writerow(filaout)                
                continue

    archivo_origen_csv = out_file
    

    archivo_origen = archivo_origen_csv
    archivo_destino = archivo_origen_csv.replace(".csv",".zip")
    contraseña = pass_file
    contraseña_bytes = contraseña.encode('utf-8') 
    with pyzipper.AESZipFile(archivo_destino, 'w', compression=zipfile.ZIP_LZMA) as zf:
        zf.setpassword(contraseña_bytes)
        zf.setencryption(pyzipper.WZ_AES, nbits=128)
        zf.write(archivo_origen)
    return archivo_destino


def get_sharepoint_context_using_app(sharepoint_url, client_id, client_secret):
    # Get sharepoint credentials
    sharepoint_url = sharepoint_url
    # Initialize the client credentials
    client_credentials = ClientCredential(client_id, client_secret)
    # create client context object
    ctx = ClientContext(sharepoint_url).with_credentials(client_credentials)
    return ctx

def send_file_to_sharepoint(sharepoint_url, client_id, client_secret, remote_path, file_name):
    logger.info(msg=f'Inicio Envia Archivo Salida Sharepoint- {sharepoint_url} : {remote_path} : {file_name}')
    target_file_url = ""
    ctx = get_sharepoint_context_using_app(sharepoint_url, client_id, client_secret)
    target_folder = ctx.web.get_folder_by_server_relative_url(remote_path)
    with open(file_name, 'rb') as content_file:
        file_content = content_file.read()
        target_file = target_folder.upload_file(file_name, file_content).execute_query()        
        content_file.close()
    site_path = sharepoint_url.split("/")[:3]
    target_file_url = site_path[0] + "//" + site_path[2] + target_file.serverRelativeUrl
    return target_file_url

def notifyResultMessage(project_id, topic_id, action_type, workflow_id, request_time, user_email, channel_notify, jsonResult, file_path=None):
    logger.info(msg=f'Inicio notifyResultMessage')
    if file_path is not None:
        message = {"action_type":action_type,
                "workflow_id":workflow_id,
                "request_time":request_time,
                "user_email":user_email,
                "file_add":"true",
                "file_path":file_path,
                "channel_notify":channel_notify,
                "payload":jsonResult,}
    else:
        message = {"action_type":action_type,
                "workflow_id":workflow_id,
                "request_time":request_time,
                "user_email":user_email,
                "file_add":"false",                
                "channel_notify":channel_notify,
                "payload":jsonResult,}
    logger.info(msg=f'notifyResultMessage msg: {message}')
    publisher = pubsub_v1.PublisherClient()
    topic_path = publisher.topic_path(project_id, topic_id)
    data_str = json.dumps(message)
    data = data_str.encode("utf-8")    
    future = publisher.publish(topic_path, data)
    logger.info(msg=f'Salida notifyResultMessage: {future.result()}')

def main():

    dbhost = os.getenv('HOSTNAME_DB')
    dbuser = os.getenv('USERNAME_DB')
    dbpass = os.getenv('PASS_DB')
    dbport = os.getenv('PORT_DB')
    dbschema = os.getenv('SCHEMA_NAME_DB')
    dbservice = os.getenv('SERVICE_NAME_DB')
    sharepoint_host = os.getenv('SHAREPOINT_HOST')
    sharepoint_user = os.getenv('SHAREPOINT_USER')
    sharepoint_pass = os.getenv('SHAREPOINT_PASS')

    projectid_topic_notify = os.getenv('RESULT_NOTIF_PROJECTID')
    topic_notify = os.getenv('RESULT_NOTIF_TOPIC')

    process_name = sys.argv[1]
    action_type = sys.argv[2]
    request_id = sys.argv[3]
    user_email = sys.argv[4]
    sharepoint_path_ori = sys.argv[5]
    sharepoint_path_dest = sys.argv[6]
    file_name_ori = sys.argv[7]
    pass_file = sys.argv[8] 
    request_time = sys.argv[9]
    workflow_id = sys.argv[10]
    channel_notify = sys.argv[11]
    
    logger.info(msg=f'Inicia Proceso RM: {process_name}')    
    try:
        local_file_path = get_rm_source_file_fromsharepoint_tocsv(sharepoint_host, sharepoint_user, sharepoint_pass, sharepoint_path_ori, file_name_ori)
    except Exception as e:
        logger.error(msg='Error en get_rm_source_file_fromsharepoint_tocsv',exc_info=True)
        jsonError = {"nombre_proceso":process_name,
                    "errors":[{"desc":"Error al transformar archivo origen get_rm_source_file_fromsharepoint_tocsv"}]}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonError)
        exit(1)
    try:        
        output_file_path = rm_complete_data(dbhost, dbuser, dbpass, dbport, dbservice, dbschema, local_file_path, pass_file)
    except Exception as e:
        logger.error(msg='Error en rm_complete_data',exc_info=True)
        jsonError = {"nombre_proceso":process_name,
                    "errors":[{"desc":"Error al completar datos rm_complete_data"}]}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonError)
        exit(1)
    try:
        output_file_share = send_file_to_sharepoint(sharepoint_host, sharepoint_user, sharepoint_pass, sharepoint_path_dest, output_file_path)
    except Exception as e:
        logger.error(msg='Error en send_file_to_sharepoint',exc_info=True)
        jsonError = {"nombre_proceso":process_name,
                    "errors":[{"desc":"Error al enviar datos a sharepoint send_file_to_sharepoint"}]}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonError)
        exit(1)
    try:
        jsonResult = {"Nombre_Proceso":process_name,
                    "Id_Solicitud":request_id,
                    "Desc.":"Se ha generado archivo de respuesta."}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonResult, file_path=output_file_share)
    except Exception as e:
        logger.error(msg='Error en generate_query',exc_info=True)
        exit(1)
    try:        
        output_file_path_rm = rm_generate_rm_file(local_file_path, pass_file)
    except Exception as e:
        logger.error(msg='Error en rm_complete_data',exc_info=True)
        jsonError = {"nombre_proceso":process_name,
                    "errors":[{"desc":"Error al completar datos rm_generate_rm_file"}]}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonError)
        exit(1)
    try:
        output_file_share_rm = send_file_to_sharepoint(sharepoint_host, sharepoint_user, sharepoint_pass, sharepoint_path_dest, output_file_path_rm)
    except Exception as e:
        logger.error(msg='Error en send_file_to_sharepoint',exc_info=True)
        jsonError = {"nombre_proceso":process_name,
                    "errors":[{"desc":"Error al enviar datos a sharepoint send_file_to_sharepoint formato rm"}]}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonError)
        exit(1)
    try:
        jsonResult = {"Nombre_Proceso":process_name,
                    "Id_Solicitud":request_id,
                    "Desc.":"Se ha generado archivo de respuesta formato RM."}
        notifyResultMessage(projectid_topic_notify, topic_notify, action_type, workflow_id, request_time, user_email, channel_notify, jsonResult, file_path=output_file_share_rm)
    except Exception as e:
        logger.error(msg='Error en generate_query',exc_info=True)
        exit(1)

if __name__ == "__main__":
    main()
